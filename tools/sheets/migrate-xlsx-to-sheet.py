#!/usr/bin/env python3
"""
One-time migration: copy every DD.MM.YYYY tab from
state/weekly-customer-sync.xlsx into the Cyvore GTM Weekly Sync Google Sheet.

After this runs successfully, the xlsx is no longer in the read path —
generate-weekly-tab.py and read-weekly-sync.py read/write the Sheet
exclusively. Archive the xlsx (this script does NOT delete or move it; the
operator does that after eyeballing the migrated Sheet).

Run order:

    1. tools/sheets/google-sheets-config.json + credentials must exist
       (see tools/sheets/google-sheets-config.template.json).
    2. The Sheet must be shared with the service-account email as Editor.
    3. python3 tools/sheets/migrate-xlsx-to-sheet.py --dry-run    # preview
    4. python3 tools/sheets/migrate-xlsx-to-sheet.py              # apply
    5. Eyeball the Sheet vs xlsx for one or two tabs to spot-check.
    6. mv state/weekly-customer-sync.xlsx state/archive/weekly-customer-sync-pre-2026-W22.xlsx

Throttling: Sheets writes are rate-limited (~60 writes/min for service
accounts). This script processes one tab per ~1.2s by default to stay well
under the limit. Override with --pause-seconds.

Behaviour:
  - Tabs already present in the Sheet are SKIPPED (idempotent re-runs).
  - Use --force to overwrite existing tabs in the Sheet (rare).
  - Use --only-week YYYY-WW to migrate a single week's tab (testing).
  - Data-validation rules ("dropdowns") in the source xlsx are translated
    into ONE_OF_LIST validations on the Sheet via SheetsClient. Inline list
    formulas (e.g. '"A,B"') are supported; range-reference list formulas
    (e.g. '=Lookups!$A$1:$A$5') are not (will be skipped with a WARN).
  - --apply-validations-only re-applies xlsx validations to existing Sheet
    tabs WITHOUT rewriting any cell values. Safe to re-run on tabs the user
    has edited post-migration.

Usage:

    python3 tools/sheets/migrate-xlsx-to-sheet.py --dry-run
    python3 tools/sheets/migrate-xlsx-to-sheet.py
    python3 tools/sheets/migrate-xlsx-to-sheet.py --only-week 2026-19
    python3 tools/sheets/migrate-xlsx-to-sheet.py --force
    python3 tools/sheets/migrate-xlsx-to-sheet.py --apply-validations-only
"""

import argparse
import importlib.util
import re
import sys
import time
from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = REPO_ROOT / "state" / "weekly-customer-sync.xlsx"

EXPECTED_HEADERS = [
    "Tier",
    "Company/Lead Name",
    "Deal/Lead Stage",
    "Status",
    "Next Step",
    "Assignee",
    "Done?",
    "moving status",
]


def die(msg, code=2):
    sys.stderr.write(f"ERROR: {msg}\n")
    sys.exit(code)


def load_sheets_client():
    spec = importlib.util.spec_from_file_location(
        "sheets_client", REPO_ROOT / "tools" / "sheets" / "sheets-client.py"
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def parse_tab_date(name):
    m = re.match(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\s*$", name)
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def iso_week_of(d):
    y, w, _ = d.isocalendar()
    return f"{y:04d}-{int(w):02d}"


_XLSX_RANGE_RE = re.compile(
    r"^\$?(?P<c1>[A-Z]+)\$?(?P<r1>\d+)(?::\$?(?P<c2>[A-Z]+)\$?(?P<r2>\d+))?$"
)


def _xlsx_col_to_index(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n  # 1-indexed


def _parse_xlsx_range(ref):
    """'A2:A54' -> (start_row=2, end_row=54, start_col=1, end_col=1) or None."""
    m = _XLSX_RANGE_RE.match(ref.strip())
    if not m:
        return None
    r1 = int(m.group("r1"))
    c1 = _xlsx_col_to_index(m.group("c1"))
    r2 = int(m.group("r2") or m.group("r1"))
    c2 = _xlsx_col_to_index(m.group("c2") or m.group("c1"))
    return {
        "start_row": min(r1, r2), "end_row": max(r1, r2),
        "start_col": min(c1, c2), "end_col": max(c1, c2),
    }


def _parse_xlsx_inline_list(formula1):
    """'"A,B"' -> ['A','B'].  Returns None if not an inline list (e.g. range ref)."""
    if formula1 is None:
        return None
    f = str(formula1).strip()
    if len(f) >= 2 and f[0] == '"' and f[-1] == '"':
        body = f[1:-1]
    elif f.startswith("'") and f.endswith("'") and len(f) >= 2:
        body = f[1:-1]
    else:
        return None
    return [v.strip() for v in body.split(",") if v.strip()]


def extract_xlsx_validations(ws, tab_name):
    """Translate openpyxl data validations to SheetsClient.apply_data_validations format.

    Only inline-list validations are emitted. Range-ref validations log a WARN
    and are skipped (not used in the Cyvore sync surface).
    """
    out = []
    for dv in getattr(ws.data_validations, "dataValidation", []) or []:
        dv_type = (dv.type or "").lower()
        if dv_type != "list":
            continue
        values = _parse_xlsx_inline_list(dv.formula1)
        if values is None:
            sys.stderr.write(
                f"[migrate] WARN {tab_name}: skipping non-inline validation "
                f"formula1={dv.formula1!r} (range refs not supported).\n"
            )
            continue
        for sqref in dv.sqref.ranges:
            rng = _parse_xlsx_range(str(sqref))
            if rng is None:
                sys.stderr.write(
                    f"[migrate] WARN {tab_name}: skipping unparseable range {sqref!r}\n"
                )
                continue
            out.append({
                "range": rng,
                "values": values,
                "strict": False,
                "show_custom_ui": True,
            })
    return out


def read_xlsx_tab(ws):
    """Return list[list[str]] for a worksheet — header row first, then data
    rows. Trailing empty rows are stripped. Cell values are stringified;
    None becomes ''."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([
            "" if c is None else (c if isinstance(c, str) else str(c))
            for c in row
        ])
    while rows and not any((c or "").strip() for c in rows[-1]):
        rows.pop()
    return rows


def main():
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    p.add_argument("--xlsx", default=str(DEFAULT_XLSX),
                   help=f"Source xlsx path. Default: {DEFAULT_XLSX.relative_to(REPO_ROOT)}")
    p.add_argument("--config", default=None,
                   help="Path to tools/sheets/google-sheets-config.json (default: standard location)")
    p.add_argument("--only-week", default=None, help="Migrate only the tab whose ISO week matches YYYY-WW")
    p.add_argument("--force", action="store_true",
                   help="Overwrite tabs that already exist in the Sheet (default: skip)")
    p.add_argument("--dry-run", action="store_true",
                   help="Print the migration plan without writing to the Sheet")
    p.add_argument("--pause-seconds", type=float, default=1.2,
                   help="Sleep between tabs to stay under Sheets API rate limit (default: 1.2)")
    p.add_argument("--apply-validations-only", action="store_true",
                   help="Re-apply xlsx data-validation rules (dropdowns) to existing Sheet "
                        "tabs WITHOUT rewriting any cell values. Safe for re-runs.")
    args = p.parse_args()

    try:
        import openpyxl
    except ImportError:
        die("openpyxl is not installed. Run: pip install -r requirements.txt")

    xlsx_path = Path(args.xlsx).expanduser()
    if not xlsx_path.is_absolute():
        xlsx_path = (REPO_ROOT / xlsx_path).resolve()
    else:
        xlsx_path = xlsx_path.resolve()
    if not xlsx_path.exists():
        die(f"xlsx not found at {xlsx_path}")

    try:
        display_path = xlsx_path.relative_to(REPO_ROOT)
    except ValueError:
        display_path = xlsx_path
    sys.stderr.write(f"[migrate] Loading xlsx: {display_path}\n")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    plan = []
    for name in wb.sheetnames:
        d = parse_tab_date(name)
        if d is None:
            sys.stderr.write(f"[migrate] Skipping non-date tab: {name!r}\n")
            continue
        if args.only_week and iso_week_of(d) != args.only_week:
            continue
        plan.append((name, d))
    plan.sort(key=lambda x: x[1])
    if not plan:
        die("No date-named tabs found in xlsx (or --only-week filter excluded them all).")

    sys.stderr.write(f"[migrate] Found {len(plan)} candidate tabs to migrate.\n")

    sheets_module = load_sheets_client()

    if args.dry_run:
        # Don't even authenticate to the Sheet on dry-run — print plan to stderr
        # so the user can pipe stdout if they want only the JSON later.
        for name, d in plan:
            ws = wb[name]
            rows = read_xlsx_tab(ws)
            sys.stderr.write(f"[migrate] [dry-run] {name} ({iso_week_of(d)}) — {len(rows)} rows\n")
        sys.stderr.write("[migrate] [dry-run] No writes performed. Re-run without --dry-run to apply.\n")
        return

    sc = sheets_module.SheetsClient(args.config)
    sys.stderr.write(f"[migrate] Connected to Sheet: {sc.spreadsheet_title} ({sc.spreadsheet_url})\n")

    existing_tabs = set(sc.list_tabs())

    # ---- Validations-only mode: skip value writes, just refresh dropdowns ----
    if args.apply_validations_only:
        applied, missing = [], []
        for name, _d in plan:
            if name not in existing_tabs:
                sys.stderr.write(
                    f"[migrate] SKIP {name}: not in Sheet (nothing to apply validations to).\n"
                )
                missing.append(name)
                continue
            ws = wb[name]
            validations = extract_xlsx_validations(ws, name)
            if not validations:
                sys.stderr.write(f"[migrate] {name}: no inline-list validations in xlsx.\n")
                continue
            sc.apply_data_validations(name, validations)
            sys.stderr.write(
                f"[migrate] APPLIED {name}: {len(validations)} validation range(s).\n"
            )
            applied.append(name)
            time.sleep(args.pause_seconds)
        sys.stderr.write("\n=== Validations-only summary ===\n")
        sys.stderr.write(f"  Applied to:        {len(applied)} tab(s) -> {applied}\n")
        if missing:
            sys.stderr.write(f"  Not in Sheet:      {missing}\n")
        sys.stderr.write(f"  Sheet:             {sc.spreadsheet_url}\n")
        return

    written, skipped, overwritten = [], [], []

    for name, d in plan:
        if name in existing_tabs and not args.force:
            sys.stderr.write(f"[migrate] SKIP {name} (already exists; --force to overwrite)\n")
            skipped.append(name)
            continue
        if name in existing_tabs and args.force:
            sys.stderr.write(f"[migrate] OVERWRITE {name} (already exists; --force)\n")
            sc.delete_tab(name)
            overwritten.append(name)

        ws = wb[name]
        rows = read_xlsx_tab(ws)
        if not rows:
            sys.stderr.write(f"[migrate] WARN {name}: empty tab in xlsx, creating empty Sheet tab anyway\n")
            sc.create_tab(name)
            written.append(name)
            time.sleep(args.pause_seconds)
            continue

        # Pad short rows to 8 columns; trim long rows to 8 columns.
        n_cols = max(len(EXPECTED_HEADERS), max(len(r) for r in rows))
        normalized = []
        for r in rows:
            padded = list(r) + [""] * (n_cols - len(r))
            normalized.append(padded[:n_cols])

        # Header sanity check (warn-only)
        header = [c.strip() for c in normalized[0][: len(EXPECTED_HEADERS)]]
        if header != EXPECTED_HEADERS:
            sys.stderr.write(
                f"[migrate] WARN {name}: header row does not match canonical 8 columns. "
                f"Got: {header}. Migrating verbatim anyway.\n"
            )

        sc.write_full_tab(name, normalized, with_header=False)
        validations = extract_xlsx_validations(ws, name)
        if validations:
            sc.apply_data_validations(name, validations)
            sys.stderr.write(
                f"[migrate] WROTE {name} ({len(normalized) - 1} data rows, "
                f"{len(validations)} validation range(s))\n"
            )
        else:
            sys.stderr.write(f"[migrate] WROTE {name} ({len(normalized) - 1} data rows)\n")
        written.append(name)
        time.sleep(args.pause_seconds)

    sys.stderr.write("\n=== Migration summary ===\n")
    sys.stderr.write(f"  Wrote new tabs:     {len(written)}\n")
    sys.stderr.write(f"  Overwrote existing: {len(overwritten)}\n")
    sys.stderr.write(f"  Skipped existing:   {len(skipped)}\n")
    sys.stderr.write(f"  Sheet:              {sc.spreadsheet_url}\n")
    sys.stderr.write("\nNext step (manual): eyeball one or two migrated tabs, then archive the xlsx:\n")
    sys.stderr.write("  mkdir -p state/archive\n")
    sys.stderr.write(
        "  mv state/weekly-customer-sync.xlsx state/archive/"
        "weekly-customer-sync-pre-2026-W22.xlsx\n"
    )


if __name__ == "__main__":
    main()
